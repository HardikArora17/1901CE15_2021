import os
import csv
from openpyxl import Workbook

i=0
name_roll_no={}
roll_dis={}
roll_subjects={}

if(not os.path.exists('tut05/output/')):
    os.mkdir('tut05/output/')
    
sub_name={}
sub_ltp={}
sub_credit={}

with open('tut05/names-roll.csv', newline='') as csvfile:
     reader = csv.DictReader(csvfile)
     for row in reader:
         name_roll_no[row['Roll']]=row['Name']
         roll_dis[row['Roll']]=row['Roll'][4:6]

with open('tut05/subjects_master.csv', newline='') as csvfile:
     reader = csv.DictReader(csvfile)
     for row in reader:
         sub_name[row['subno']]=row['subname']
         sub_ltp[row['subno']]=row['ltp']
         sub_credit[row['subno']]=row['crd']
         
         
with open('tut05/grades.csv', newline='') as csvfile:
     reader = csv.DictReader(csvfile)
     for row in reader:
         roll_no=row['Roll']
         sem=int(row['Sem'])
         sub_code=row['SubCode']
         sub_type=row['Sub_Type']
         credit=int(row['Credit'])
         grade=row['Grade']
        
         if roll_no not in roll_subjects.keys(): 
            roll_subjects[roll_no]={}
            roll_subjects[roll_no][sem]=[]
            roll_subjects[roll_no][sem].append((credit,grade,sub_code,sub_type))
         else:
            if(sem not in roll_subjects[roll_no].keys()):
                roll_subjects[roll_no][sem]=[]
                roll_subjects[roll_no][sem].append((credit,grade,sub_code,sub_type))
            else:
                roll_subjects[roll_no][sem].append((credit,grade,sub_code,sub_type))


roll_result={}
roll_sem_wise={}

d={'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0}


for r_no,s in roll_subjects.items(): 
    for k,v in s.items():
        
        credits=[v1[0] for v1 in v]
        grade=  [v1[1] for v1 in v]
        sub_code=[v1[2] for v1 in v]
        sub_type=[v1[3] for v1 in v]
        
        total=sum(credits)
        p=0
        for i in range(len(grade)):
            grade[i]=grade[i].strip()
            
            if(grade[i].endswith('*')):
                grade[i]=grade[i][:-1]
                
            p+=d[grade[i]]*credits[i]
            
        spi=p/total
        spi=round(spi,2)
        if(r_no not in roll_result.keys()):
            roll_result[r_no]=[]
            roll_sem_wise[r_no]=[]
            
        roll_sem_wise[r_no].append((k,sub_code,sub_type,grade))  
        roll_result[r_no].append((k,total,spi))
        

for r_no,s in roll_result.items():
    wb1=Workbook()
    filepath="tut05/output/"+r_no+".xlsx"     
    
    wb=wb1.create_sheet('Overall',0)
    wb.append(["Roll No",r_no])
    wb.append(["Name of Student",name_roll_no[r_no]])
    wb.append(["Discipline",roll_dis[r_no]])
    wb.append(["Semester No.",1,2,3,4,5,6,7,8])
    
    spi=[item[2] for item in s]
    credits_sem_wise=[item[1] for item in s]
    CPI=[]
    total_credits_sem=[]
    t=0
    c=0
    for i in range(len(credits_sem_wise)):
        t+=credits_sem_wise[i]
        c+=credits_sem_wise[i]*spi[i]
        
        temp=c/t
        temp=round(temp,2)
        CPI.append(temp)
        total_credits_sem.append(t)
             
    columns=["Semester wise Credit Taken","SPI","Total Credits Taken","CPI"]
    tags={0:spi,1:credits_sem_wise,2:total_credits_sem,3:CPI}
    
    for i,name in enumerate(columns):
        a=[name]+tags[i]
        wb.append(a)
    
    
    for i in range(len(roll_sem_wise[r_no])):
        
        wb2=wb1.create_sheet('SEM '+str(i+1),i+1)
        
        fixed_line=["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"]
        wb2.append(fixed_line)
        
        for j in range(len(roll_sem_wise[r_no][i][2])):
            kl=roll_sem_wise[r_no][i]
            wb2.append([j+1,kl[1][j],sub_name[kl[1][j]],sub_ltp[kl[1][j]],sub_credit[kl[1][j]],kl[2][j],kl[3][j]])


    wb1.save(filepath)
    



    


            
                
        
        
        
                
 
                
