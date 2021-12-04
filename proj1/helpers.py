import streamlit as st

import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

from stqdm import stqdm

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils import get_styles
import pandas as pd

bold_, thin, heading_, correct_style, wrong_style, neut_style, true_style = get_styles()

def insert_info(sheet, name, rolln, absent_status):

    ''' INSERT STUDENT INFO IN TABLE1 '''

    # Name
    sheet.cell(row = 6, column = 1).value = 'Name: '
    sheet.cell(row = 6, column = 1).font = Font(name='Century', size = 12)
    sheet.cell(row = 6, column = 1).alignment = Alignment(horizontal='right')
    sheet.cell(row = 6, column = 2).value = name
    sheet.cell(row = 6, column = 2).font = Font(name='Century', size = 12, bold=True)
    
    # Roll No.
    sheet.cell(row = 7, column = 1).value = 'Roll Numer: '
    sheet.cell(row = 7, column = 1).font = Font(name='Century', size = 12)
    sheet.cell(row = 7, column = 1).alignment = Alignment(horizontal='right')
    sheet.cell(row = 7, column = 2).value = rolln
    sheet.cell(row = 7, column = 2).font = Font(name='Century', size = 12, bold=True)

    # Exam
    sheet.cell(row = 6, column = 4).value = 'Exam: '
    if(absent_status==1):
        sheet.cell(row = 7, column = 4).value = 'Absent '
        sheet.cell(row = 7, column = 4).font = Font(name='Century', size = 12, bold=True,color='00ff0000')

    sheet.cell(row = 6, column = 4).font = Font(name='Century', size = 12)
    sheet.cell(row = 6, column = 4).alignment = Alignment(horizontal='right')
    sheet.cell(row = 6, column = 5).value = 'quiz'
    sheet.cell(row = 6, column = 5).font = Font(name='Century', size = 12, bold=True)

    return sheet


def add_borders(sheet, n_qs):

    ''' ADD BORDERS IN BOTH THE TABLES '''
    for i in range(9,13):
        for j in range(1,6):
            cell_ = sheet.cell(row = i, column = j)
            thin = Side(border_style="thin", color="000000")
            cell_.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if n_qs <= 25:
        for i in range(15, 15+n_qs+1):
            for j in range(1,3):
                cell_ = sheet.cell(row = i, column = j)
                thin = Side(border_style="thin", color="000000")
                cell_.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    else:
        for i in range(15, 41):
            for j in range(1,3):
                cell_ = sheet.cell(row = i, column = j)
                thin = Side(border_style="thin", color="000000")
                cell_.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for i in range(15, 15+n_qs-25+1):
            for j in range(4,6):
                cell_ = sheet.cell(row = i, column = j)
                thin = Side(border_style="thin", color="000000")
                cell_.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    return sheet

def add_headings(sheet, n_qs):

    ''' ADD HEADINGS IN BOTH THE TABLES '''
    row1 = ['Right','Wrong','Not Attempt','Max']
    for c in range(4):
        sheet.cell(row = 9, column = c+2).value = row1[c]
        sheet.cell(row = 9, column = c+2).style = heading_

    col1 = ['No.', 'Marking', 'Total']
    for c in range(3):
        sheet.cell(row = 10+c, column = 1).value = col1[c]
        sheet.cell(row = 10+c, column = 1).style = heading_

    row2 = ['Student Ans', 'Correct Ans']
    for c in range(2):
        sheet.cell(row = 15, column = c+1).value = row2[c]
        sheet.cell(row = 15, column = c+1).style = heading_
    
    if n_qs >25:
        for c in range(2):
            sheet.cell(row = 15, column = c+4).value = row2[c]
            sheet.cell(row = 15, column = c+4).style = heading_

    return sheet

def fill_table_1(sheet, right, wrong, unatmp, right_marks, wrong_marks, neut_marks):

    ''' FILL THE FIRST TABLE '''
    col1 = (right, right_marks, right*right_marks) # 1st column
    for c in range(3):
        sheet.cell(row = 10+c, column = 2).value = col1[c]
        sheet.cell(row = 10+c, column = 2).style = correct_style

    col2 = (wrong, wrong_marks, wrong*wrong_marks) # 2nd column
    for c in range(3):
        sheet.cell(row = 10+c, column = 3).value = col2[c]
        sheet.cell(row = 10+c, column = 3).style = wrong_style

    col3 = (unatmp, neut_marks, unatmp*neut_marks) # 3rd column
    for c in range(3):
        if c==2 and col3[c]==0:
            continue
        else:
            sheet.cell(row = 10+c, column = 4).value = col3[c]
            sheet.cell(row = 10+c, column = 4).style = neut_style

    # Total row
    sheet.cell(row = 10, column = 5).value = right+wrong+unatmp
    sheet.cell(row = 10, column = 5).style = neut_style

    sheet.cell(row = 12, column = 5).value = '{}/{}'.format((right*right_marks+wrong*wrong_marks+unatmp*neut_marks), ((right+wrong+unatmp)*right_marks))
    sheet.cell(row = 12, column = 5).style = true_style

    return sheet

def gen_roll_wise(responses, master_roll, right_marks, wrong_marks, neut_marks): # If generate roll no. wise marksheets button is clicked

    ''' GENERATES ROLL-WISE EXCEL MARKSHEETS '''

    correct_idx = responses[responses['Roll Number'].isin(['ANSWER'])].index[0]
    correct_row = responses.iloc[correct_idx].to_list()
    ANS_idx = correct_row.index('ANSWER')+1
    n_qs = responses.shape[1] - ANS_idx
    correct_resps = correct_row[ANS_idx:]
    

    # Preprocess the data
    responses = responses.drop_duplicates(subset='Roll Number')
    responses = responses.fillna('None')
    master_roll = master_roll.drop_duplicates(subset='roll')

    net_scores = [] # Scores after negative marking 
    absent_rows = [] # Rows of the absentees (to append at the end of code, in the concise marksheet)

    # Iterate in responses
    x={}
    
    for i in stqdm(range(responses.shape[0])):
        resp = responses.loc[i].to_list() # student's responses to the google form
        rolln = resp[6] # Student's roll number
        x[rolln]=resp
        
        
    for i in stqdm(range(master_roll.shape[0])):

        # Absentee Check
        absent_status=0
        rolln, name = master_roll.loc[i].to_list()
        
        # if rolln not in responses.values:
        #     wb = openpyxl.Workbook()
        #     sheet = wb.worksheets[0]
        #     sheet.title = 'quiz'
        #     wb.save('sample_output/marksheet/{}.xlsx'.format(rolln))
        #     absent = [None]*2+['ABSENT', name]+[None]*2+['ABSENT', rolln]+[None]*n_qs
        #     absent_rows.append(absent)
        #     continue
            
        # Open workbook
        wb = openpyxl.Workbook()
        sheet = wb.worksheets[0]
        sheet.title = 'quiz'

        # Mark Sheet - Heading
        sheet.cell(row = 5, column = 3).value = 'Mark Sheet'
        sheet.merge_cells('A5:E5')
        sheet.cell(row = 5, column = 3).font = Font(name='Century', size = 18, bold = True, underline='double')
        sheet.cell(row = 5, column = 3).alignment = Alignment(horizontal='center')

        # Expanding column widths
        cols = ['A','B','C','D','E']
        for col in cols:
            sheet.column_dimensions[col].width = 17.7

        if rolln not in x.keys():
            stud_resps=['']*len(correct_resps)
            absent_status=1
            
        else:
            resp=x[rolln]
            name = resp[3] # Student's name
            rolln = resp[6] # Student's roll number
            stud_resps = resp[ANS_idx:] # student's responses to exam questions
            
        
        # Inserting IITP logo
        img = openpyxl.drawing.image.Image('iitp_logo.png')
        img.anchor = 'A1'
        img.width = 123*5.03
        img.height = 20*4.45
        sheet.add_image(img)

        # Calling the designated functions
        sheet = insert_info(sheet, name, rolln,absent_status)
        sheet = add_borders(sheet, n_qs)
        sheet = add_headings(sheet, n_qs)
        
        
        # Initialize the variables used
        right = 0
        unatmp = 0
        wrong = 0
        stud_row = 1
        ans_row = 2
        row_offset = 15

        # Iterate in Student's question responses, and fill table 2
        for idx, opts in enumerate(zip(stud_resps, correct_resps)):
            qno = idx+1
            opt, ans = opts
            if qno==26: # adjust the offset of table 2 is the no. of qs exceed 25 (i.e. if the excel file reaches the 40th row)
                stud_row = 4
                ans_row = 5
                row_offset = -10
                
            sheet.cell(row = qno+row_offset, column = ans_row).value = ans
            sheet.cell(row = qno+row_offset, column = ans_row).style = true_style
            sheet.cell(row = qno+row_offset, column = stud_row).value = opt
            if opt == ans:
                sheet.cell(row = qno+row_offset, column = stud_row).style = correct_style
                right+=1
            elif absent_status or opt == 'None':
                sheet.cell(row = qno+row_offset, column = stud_row).value = None
                # sheet.cell(row = qno+row_offset, column = stud_row).style = neut_style
                unatmp+=1
            else:
                sheet.cell(row = qno+row_offset, column = stud_row).style = wrong_style
                wrong+=1

        # Fill table 1, append net_score
        sheet = fill_table_1(sheet, right, wrong, unatmp, right_marks, wrong_marks, neut_marks)
        net_scores.append('{} / {}'.format((right*right_marks+wrong*wrong_marks+unatmp*neut_marks), ((right+wrong+unatmp)*right_marks)))
        wb.save('marksheet/{}.xlsx'.format(rolln))

    return responses, net_scores, absent_rows

def get_net_scores_and_absentees(responses, master_roll, right_marks, wrong_marks, neut_marks):

    # Find the answer rows and determine the index of ANSWER row, calculate the number of questions
    correct_idx = responses[responses['Roll Number'].isin(['ANSWER'])].index[0]
    correct_row = responses.iloc[correct_idx].to_list()
    ANS_idx = correct_row.index('ANSWER')+1
    n_qs = responses.shape[1] - ANS_idx
    correct_resps = correct_row[ANS_idx:]

    # Preprocess the data
    responses = responses.drop_duplicates(subset='Roll Number')
    responses = responses.fillna('None')
    master_roll = master_roll.drop_duplicates(subset='roll')
    
    net_scores = [] # Scores after negative marking 
    absent_rows = [] # Rows of the absentees (to append at the end of code, in the concise marksheet)
    google_scores=[]
    statusAns=[]
    
    for i in stqdm(range(master_roll.shape[0])):

        # Absentee Check
        rolln, name = master_roll.loc[i].to_list()
        if rolln not in responses.values:
            absent = [None]*2+['ABSENT', name]+[None]*2+['ABSENT', rolln]+[None]*(n_qs+1)
            absent_rows.append(absent)
            continue

        resp = responses.loc[i].to_list() # student's responses to the google form
        name = resp[3] # Student's name
        rolln = resp[6] # Student's roll number
        stud_resps = resp[ANS_idx:] # student's responses to exam questions

        # Initialize the variables used
        right = 0
        unatmp = 0
        wrong = 0

        # Iterate in Student's question responses
        for idx, opts in enumerate(zip(stud_resps, correct_resps)):
            qno = idx+1
            opt, ans = opts
            if opt == ans:
                right+=1
            elif opt == 'None':
                unatmp+=1
            else:
                wrong+=1

        # Append net_score
        net_scores.append('{} / {}'.format(round((right*right_marks+wrong*wrong_marks+unatmp*neut_marks),2), round(((right+wrong+unatmp)*right_marks),2)))
        google_scores.append('{} / {}'.format(round((right*right_marks+unatmp*neut_marks),2), round(((right+wrong+unatmp)*right_marks),2)))
        statusAns.append('{},{},{}'.format(right,wrong,unatmp))
        
    responses.loc[:,['Score']] = pd.Series(google_scores)
    responses.insert(len(responses.columns),'statusAns',statusAns)
    
    return responses, net_scores, absent_rows

def gen_concise(responses, net_scores, absent_rows):  # If generate concise marksheet button is clicked

    ''' GENERATE CONCISE MARKSHEET '''

    responses.rename(columns = {'Score':'Google_Score'}, inplace = True) # rename score column
    responses.insert(loc=6, column='Score_After_Negative', value=net_scores) # insert new scores after negative marking
    
    # replace 'None' with empty strings
    mask = responses.applymap(lambda x: x == 'None')
    cols = responses.columns[(mask).any()]
    
    for col in responses[cols]:
        responses.loc[mask[col], col] = ''

    # append absentee rows and save the concise marksheet
    for absent_row in absent_rows:
        responses.loc[len(responses)] = absent_row
        
    responses.to_csv('marksheet/concise_marksheet.csv', index=False)

    return responses

def get_concise(responses, net_scores, absent_rows):  # If generate concise marksheet button is clicked

    ''' GENERATE CONCISE MARKSHEET '''

    responses.rename(columns = {'Score':'Google_Score'}, inplace = True) # rename score column
    responses.insert(loc=6, column='Score_After_Negative', value=net_scores) # insert new scores after negative marking
    
    # replace 'None' with empty strings
    mask = responses.applymap(lambda x: x == 'None')
    cols = responses.columns[(mask).any()]
    for col in responses[cols]:
        responses.loc[mask[col], col] = ''

    # append absentee rows and save the concise marksheet
    for absent_row in absent_rows:
        responses.loc[len(responses)] = absent_row
    responses.to_csv('marksheet/concise_marksheet.csv', index=False)

    return responses

def send_mails(responses,flag_file):

    EMAIL_ADDRESS = 'zeus200no@gmail.com'
    EMAIL_PASSWORD = 'gelt ltax ixyk vqho'

    responses = responses[['Email address', 'IITP webmail', 'Roll Number']]

    for idx in stqdm(range(responses.shape[0])):

        info = responses.loc[idx].to_list()
        rolln = info[2]
        emails = info[0:2]
        #st.write(emails)
        
        if rolln=='ANSWER':
            continue

        if emails != []:
            msg = MIMEMultipart()
            msg['From'] = EMAIL_ADDRESS
            msg['To'] = ', '.join(emails)
            msg['Date'] = formatdate(localtime = True)
            msg['Subject'] = 'Quiz Results'
            msg.attach(MIMEText('Kindly find the results of previous quiz in the attached file'))

            part = MIMEBase('application', "octet-stream")
            part.set_payload(open("marksheet/{}.xlsx".format(rolln), "rb").read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename="{}.xlsx"'.format(rolln))
            msg.attach(part)
            RECEIVER_EMAIL=emails[1]
            
            with smtplib.SMTP('smtp.gmail.com', 587) as smtp:

                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()

                smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
                smtp.sendmail(EMAIL_ADDRESS, RECEIVER_EMAIL, msg.as_string())
                
        else:
            continue
        
        
    
    return 0